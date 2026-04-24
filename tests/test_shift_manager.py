from __future__ import annotations

import csv
import sqlite3
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

from openpyxl import load_workbook

from shift_manager import (
    MasterLogWriter,
    ShiftManager,
    create_shift_manager,
    get_shift_id,
    get_shift_time_range,
    normalize_event,
    parse_log_line,
)


UTC = timezone.utc
LOCAL_TZ = datetime.now().astimezone().tzinfo or UTC


def test_get_shift_id() -> None:
    assert get_shift_id(datetime(2026, 3, 18, 1, 0, tzinfo=UTC)) == 1
    assert get_shift_id(datetime(2026, 3, 18, 9, 0, tzinfo=UTC)) == 2
    assert get_shift_id(datetime(2026, 3, 18, 18, 0, tzinfo=UTC)) == 3


def test_get_shift_time_range() -> None:
    target = date(2026, 3, 18)
    start1, end1 = get_shift_time_range(1, target)
    start2, end2 = get_shift_time_range(2, target)
    start3, end3 = get_shift_time_range(3, target)
    assert start1.hour == 0 and end1.hour == 8
    assert start2.hour == 8 and end2.hour == 16
    assert start3.hour == 16 and end3.day == 19


def test_master_log_writer_round_trip(tmp_path: Path) -> None:
    writer = MasterLogWriter(tmp_path / "user_log.txt")
    writer.write(
        "APP_USAGE",
        {
            "ended_at": "2026-03-18T10:00:00Z",
            "duration_seconds": 60,
            "app_name": "Code",
            "window_title": "main.py - VS Code",
        },
    )
    entries = writer.read_all()
    assert len(entries) == 1
    assert entries[0].category == "APP"
    assert entries[0].item == "Code"
    assert entries[0].duration_seconds == 60


def test_pipeline_db_is_idempotent(tmp_path: Path) -> None:
    pipeline = create_shift_manager(tmp_path, log_path=tmp_path / "user_log.txt", database_path=tmp_path / "user_log.db")
    payload = {
        "ended_at": "2026-03-18T10:00:00Z",
        "duration_seconds": 60,
        "app_name": "Code",
        "window_title": "main.py - VS Code",
    }
    pipeline.log("APP_USAGE", payload)
    pipeline.log("APP_USAGE", payload)
    assert pipeline.database.count_events() == 1


def test_daily_workbook_generation(tmp_path: Path) -> None:
    pipeline = create_shift_manager(tmp_path, log_path=tmp_path / "user_log.txt", database_path=tmp_path / "user_log.db", export_dir=tmp_path)
    pipeline.log(
        "APP_USAGE",
        {"ended_at": "2026-03-18T06:00:00Z", "duration_seconds": 120, "app_name": "Code", "window_title": "repo"},
    )
    pipeline.log(
        "WEBSITE_USAGE",
        {"ended_at": "2026-03-18T10:00:00Z", "duration_seconds": 240, "domain": "github.com", "url": "https://github.com"},
    )
    pipeline.log(
        "MEDIA_PLAYBACK",
        {"ended_at": "2026-03-18T18:00:00Z", "duration_seconds": 300, "source_app": "Spotify", "title": "Song A", "artist": "Artist A"},
    )

    shift1_csv = pipeline.generate_shift_csv(date(2026, 3, 18), 1)
    shift2_csv = pipeline.generate_shift_csv(date(2026, 3, 18), 2)
    shift3_csv = pipeline.generate_shift_csv(date(2026, 3, 18), 3)
    daily_csv = pipeline.generate_daily_csv(date(2026, 3, 18))
    workbook_path = pipeline.generate_daily_workbook(date(2026, 3, 18))
    assert shift1_csv.exists()
    assert shift2_csv.exists()
    assert shift3_csv.exists()
    assert daily_csv.exists()
    assert workbook_path.exists()

    with daily_csv.open("r", encoding="utf-8", newline="") as handle:
        rows = list(csv.reader(handle))
    assert rows[0][:5] == ["timestamp", "category", "item", "duration_seconds", "duration_readable"]
    assert len(rows) == 4

    workbook = load_workbook(workbook_path)
    assert workbook.sheetnames == [
        "Usage_Log",
        "Browser_Log",
        "Media_Log",
        "Shift1",
        "Shift2",
        "Shift3",
        "Summary",
    ]

    usage_rows = list(workbook["Usage_Log"].iter_rows(values_only=True))
    assert usage_rows[0][:5] == ("timestamp", "category", "item", "duration_seconds", "duration_readable")
    assert usage_rows[1][1] == "APP"

    summary_rows = list(workbook["Summary"].iter_rows(values_only=True))
    metrics = {row[0]: row[1] for row in summary_rows[1:]}
    assert metrics["total_time_seconds"] == 660


def test_consolidate_legacy_logs_deduplicates_text_and_csv(tmp_path: Path) -> None:
    legacy_text = tmp_path / "usage_log.txt"
    legacy_csv_dir = tmp_path / "table_logs"
    legacy_csv_dir.mkdir(parents=True, exist_ok=True)

    entry = normalize_event(
        "APP_USAGE",
        {
            "ended_at": "2026-03-18T10:00:00Z",
            "duration_seconds": 60,
            "app_name": "Code",
            "window_title": "repo",
        },
    )
    legacy_text.write_text(entry.to_line() + "\n", encoding="utf-8")

    with (legacy_csv_dir / "app_usage.csv").open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "logged_at",
                "started_at",
                "ended_at",
                "duration_seconds",
                "process_name",
                "app_name",
                "window_title",
            ],
        )
        writer.writeheader()
        writer.writerow(
            {
                "logged_at": "2026-03-18T10:00:00Z",
                "started_at": "2026-03-18T09:59:00Z",
                "ended_at": "2026-03-18T10:00:00Z",
                "duration_seconds": 60,
                "process_name": "Code.exe",
                "app_name": "Code",
                "window_title": "repo",
            }
        )

    pipeline = create_shift_manager(tmp_path, log_path=tmp_path / "user_log.txt", database_path=tmp_path / "user_log.db")
    result = pipeline.consolidate_legacy_logs(delete_old_files=False)
    assert result["merged_records"] == 1
    assert pipeline.database.count_events() == 1
    assert len(pipeline.master_log.read_all()) == 1


def test_consolidate_legacy_export_csvs(tmp_path: Path) -> None:
    legacy_daily_dir = tmp_path / "daily_csvs"
    legacy_daily_dir.mkdir(parents=True, exist_ok=True)

    with (legacy_daily_dir / "daily_log_2026-03-18.csv").open(
        "w", encoding="utf-8", newline=""
    ) as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "timestamp",
                "category",
                "item",
                "duration_seconds",
                "duration_readable",
                "source",
                "event_type",
                "window",
                "url",
            ],
        )
        writer.writeheader()
        writer.writerow(
            {
                "timestamp": "2026-03-18 10:00:00.000000",
                "category": "APP",
                "item": "code.exe",
                "duration_seconds": "60",
                "duration_readable": "1m 0s",
                "source": "Code",
                "event_type": "APP_USAGE",
                "window": "repo",
                "url": "",
            }
        )

    pipeline = create_shift_manager(
        tmp_path,
        log_path=tmp_path / "user_log.txt",
        database_path=tmp_path / "user_log.db",
    )
    result = pipeline.consolidate_legacy_logs(delete_old_files=False)

    assert result["merged_records"] == 1
    assert pipeline.database.count_events() == 1
    assert len(pipeline.master_log.read_all()) == 1


def test_db_failure_is_healed_from_log(tmp_path: Path, monkeypatch) -> None:
    pipeline = create_shift_manager(tmp_path, log_path=tmp_path / "user_log.txt", database_path=tmp_path / "user_log.db")
    real_upsert = pipeline.database.upsert_entry
    state = {"calls": 0}

    def flaky_upsert(entry, *, log_written: bool) -> None:
        state["calls"] += 1
        if state["calls"] == 1:
            raise sqlite3.OperationalError("db unavailable")
        real_upsert(entry, log_written=log_written)

    monkeypatch.setattr(pipeline.database, "upsert_entry", flaky_upsert)
    pipeline.log(
        "APP_USAGE",
        {"ended_at": "2026-03-18T10:00:00Z", "duration_seconds": 60, "app_name": "Code", "window_title": "repo"},
    )
    assert pipeline.database.count_events() == 0

    monkeypatch.setattr(pipeline.database, "upsert_entry", real_upsert)
    synced = pipeline.sync_master_log_to_db()
    assert synced >= 1
    assert pipeline.database.count_events() == 1


def test_file_write_failure_is_healed_from_db(tmp_path: Path, monkeypatch) -> None:
    pipeline = create_shift_manager(tmp_path, log_path=tmp_path / "user_log.txt", database_path=tmp_path / "user_log.db")
    real_append = pipeline.master_log.append_entry
    state = {"calls": 0}

    def flaky_append(entry):
        state["calls"] += 1
        if state["calls"] == 1:
            raise OSError("disk busy")
        return real_append(entry)

    monkeypatch.setattr(pipeline.master_log, "append_entry", flaky_append)
    pipeline.log(
        "WEBSITE_USAGE",
        {"ended_at": "2026-03-18T11:00:00Z", "duration_seconds": 120, "domain": "github.com", "url": "https://github.com"},
    )
    assert pipeline.database.count_events() == 1
    assert pipeline.log_path.exists() is False or len(pipeline.master_log.read_all()) == 0

    monkeypatch.setattr(pipeline.master_log, "append_entry", real_append)
    repaired = pipeline.repair_master_log_from_db()
    assert repaired == 1
    lines = pipeline.master_log.read_all()
    assert len(lines) == 1
    assert lines[0].item == "github.com"


def test_weekly_schedule_uses_friday_volume(tmp_path: Path) -> None:
    pipeline = create_shift_manager(tmp_path, log_path=tmp_path / "user_log.txt", database_path=tmp_path / "user_log.db")
    week_start = date(2026, 3, 16)
    friday = week_start + timedelta(days=4)
    pipeline.log(
        "APP_USAGE",
        {
            "ended_at": f"{friday.isoformat()}T10:00:00+00:00",
            "duration_seconds": 5 * 3600,
            "app_name": "Code",
            "window_title": "repo",
        },
    )
    assert pipeline.determine_weekly_update_day(week_start) == 5


def test_process_scheduled_tasks_generates_shift_and_daily_csv(tmp_path: Path) -> None:
    pipeline = create_shift_manager(
        tmp_path,
        log_path=tmp_path / "user_log.txt",
        database_path=tmp_path / "user_log.db",
        export_dir=tmp_path,
    )
    pipeline.log(
        "APP_USAGE",
        {
            "ended_at": "2026-03-18T02:00:00Z",
            "duration_seconds": 60,
            "app_name": "Code",
            "window_title": "repo",
        },
    )
    pipeline.log(
        "WEBSITE_USAGE",
        {
            "ended_at": "2026-03-18T10:00:00Z",
            "duration_seconds": 120,
            "domain": "github.com",
            "url": "https://github.com",
        },
    )
    result = pipeline.process_scheduled_tasks(datetime(2026, 3, 19, 1, 0, tzinfo=UTC))
    assert result["shift_csvs"]
    assert result["daily_csvs"]
    assert result["daily_workbooks"]
    assert pipeline.get_shift_csv_path(date(2026, 3, 18), 3).exists()
    assert pipeline.get_daily_csv_path(date(2026, 3, 18)).exists()


def test_process_scheduled_tasks_does_not_export_incomplete_day(tmp_path: Path) -> None:
    pipeline = create_shift_manager(
        tmp_path,
        log_path=tmp_path / "user_log.txt",
        database_path=tmp_path / "user_log.db",
        export_dir=tmp_path,
    )
    pipeline.log(
        "APP_USAGE",
        {
            "ended_at": "2026-03-18T14:00:00Z",
            "duration_seconds": 60,
            "app_name": "Code",
            "window_title": "repo",
        },
    )
    result = pipeline.process_scheduled_tasks(datetime(2026, 3, 18, 18, 0, tzinfo=UTC))
    assert result["daily_csvs"] == []
    assert result["daily_workbooks"] == []
    assert not pipeline.get_daily_csv_path(date(2026, 3, 18)).exists()


def test_process_scheduled_tasks_clamps_future_daily_state(tmp_path: Path) -> None:
    pipeline = create_shift_manager(
        tmp_path,
        log_path=tmp_path / "user_log.txt",
        database_path=tmp_path / "user_log.db",
        export_dir=tmp_path,
    )
    pipeline.log(
        "APP_USAGE",
        {
            "ended_at": "2026-03-18T10:00:00Z",
            "duration_seconds": 60,
            "app_name": "Code",
            "window_title": "repo",
        },
    )
    pipeline.database.set_state("last_daily_export_date", "2026-03-18")
    result = pipeline.process_scheduled_tasks(datetime(2026, 3, 18, 18, 0, tzinfo=UTC))
    assert result["daily_csvs"] == []
    assert pipeline.database.get_state("last_daily_export_date") == "2026-03-17"


def test_prune_before_keeps_cutoff_day_and_resets_export_state(tmp_path: Path) -> None:
    pipeline = create_shift_manager(
        tmp_path,
        log_path=tmp_path / "user_log.txt",
        database_path=tmp_path / "user_log.db",
        export_dir=tmp_path,
    )
    cutoff_local = datetime(2026, 3, 31, 0, 0, tzinfo=LOCAL_TZ)

    pipeline.log(
        "APP_USAGE",
        {
            "duration_seconds": 60,
            "app_name": "Code",
            "window_title": "old entry",
        },
        logged_at=cutoff_local - timedelta(seconds=1),
    )
    pipeline.log(
        "APP_USAGE",
        {
            "duration_seconds": 60,
            "app_name": "Code",
            "window_title": "new entry",
        },
        logged_at=cutoff_local + timedelta(seconds=1),
    )
    pipeline.database.set_state("last_shift_export_key", "2026-03-30-shift3")
    pipeline.database.set_state("last_daily_export_date", "2026-03-30")
    pipeline.database.record_shift_run(date(2026, 3, 30), 3, [])
    pipeline.database.record_daily_export(
        date(2026, 3, 30),
        tmp_path / "daily_log_2026-03-30.xlsx",
    )

    summary = pipeline.prune_before(cutoff_local)

    assert summary["removed_log_entries"] == 1
    assert summary["removed_events"] == 1
    assert summary["remaining_events"] == 1
    assert len(pipeline.master_log.read_all()) == 1
    assert pipeline.master_log.read_all()[0].window == "new entry"
    assert pipeline.database.get_state("last_log_offset") == str(
        pipeline.master_log.get_size()
    )
    assert pipeline.database.get_state("last_shift_export_key") == ""
    assert pipeline.database.get_state("last_daily_export_date") == ""

    with sqlite3.connect(tmp_path / "user_log.db") as conn:
        shift_runs = conn.execute("SELECT COUNT(*) FROM shift_runs").fetchone()[0]
        daily_exports = conn.execute("SELECT COUNT(*) FROM daily_exports").fetchone()[0]

    assert shift_runs == 0
    assert daily_exports == 0


def test_parse_legacy_and_structured_lines() -> None:
    structured = '[2026-03-18 10:00:00] | APP | Code | {"event_type":"APP_USAGE","duration_seconds":60,"window_title":"repo"}'
    legacy = '{"logged_at":"2026-03-18T10:00:00Z","event_type":"APP_USAGE","data":{"duration_seconds":60,"app_name":"Code","window_title":"repo"}}'
    assert parse_log_line(structured) is not None
    assert parse_log_line(legacy) is not None
